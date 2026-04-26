"""
dal/parsers/eventlink.py — Eventlink officiating payment history parser.
"""

import csv
import io
import json
import logging
import sqlite3
import re
from datetime import datetime, timezone

from dal.parsers.base import DocumentParser, ParseResult

log = logging.getLogger("sentry.parsers.eventlink")


class EventlinkParser(DocumentParser):
    @property
    def parser_type(self) -> str:
        return "eventlink"

    def can_parse(self, filename: str, content_bytes: bytes) -> bool:
        """Detect Eventlink exports by filename pattern or content markers."""
        name = filename.lower()
        if "eventlink" in name or "payment_history" in name:
            return True
            
        try:
            # Check content for Eventlink headers
            text = content_bytes.decode("utf-8", errors="ignore")[:2000]
            markers = ["Game Date", "Pay Rate", "Assignor", "Official", "Pay Date"]
            return sum(1 for m in markers if m.lower() in text.lower()) >= 2
        except Exception:
            try:
                # Might be binary XLSX
                if content_bytes.startswith(b"PK"):
                    if "payment" in name or "export" in name:
                        # Need openpyxl to be sure, but filename hinting + PK ZIP header is a good proxy 
                        # for Excel if we assume reasonable constraint.
                        return True
            except Exception:
                pass
            return False

    def parse(self, content_bytes: bytes) -> ParseResult:
        """Parse the export file into structured payment records."""
        # Try XLSX first, fallback to CSV
        try:
            if content_bytes.startswith(b"PK\x03\x04"):
                return self._parse_xlsx(content_bytes)
        except Exception:
            pass
            
        return self._parse_csv(content_bytes)

    def _parse_xlsx(self, file_bytes: bytes) -> ParseResult:
        """Parse XLSX format using openpyxl."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required to parse Eventlink XLSX files. Please install it.")
            
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        
        headers = []
        payments = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(cell).strip().lower() if cell is not None else "" for cell in row]
                continue
            
            row_dict = dict(zip(headers, row))
            if any(v is not None for v in row_dict.values()):
                payments.append(self._process_row(row_dict))
                
        return self._build_result(payments)

    def _parse_csv(self, file_bytes: bytes) -> ParseResult:
        """Parse CSV format."""
        content_str = file_bytes.decode("utf-8", errors="ignore")
        # Try standard parsing first
        try:
            reader = csv.DictReader(content_str.splitlines())
            rows = list(reader)
        except Exception:
            # Fallback for weird newlines in unquoted fields
            reader = csv.DictReader(content_str.splitlines(), quoting=csv.QUOTE_NONE)
            rows = list(reader)
        
        payments = []
        for row in rows:
            clean_row = {k.strip().lower(): v for k, v in row.items() if k}
            if any(clean_row.values()):
                payments.append(self._process_row(clean_row))
                
        return self._build_result(payments)
        
    def _process_row(self, row: dict) -> dict:
        """Extract needed fields flexibly given varying column names."""
        pay = {}
        # Date processing
        game_date = row.get("game date") or row.get("date") or ""
        pay_date = row.get("pay date") or row.get("payment date") or row.get("date") or ""
        
        # Format string dates to YYYY-MM-DD
        pay["game_date"] = self._format_date(game_date)
        pay["pay_date"] = self._format_date(pay_date) or pay["game_date"]
        
        # Amount
        amt_str = str(row.get("amount", row.get("total", row.get("pay rate", "0"))))
        amt_str = amt_str.replace('$', '').replace(',', '').strip()
        try:
            pay["amount"] = float(amt_str)
        except ValueError:
            pay["amount"] = 0.0
            
        # Details
        pay["sport"] = row.get("sport", row.get("event", ""))
        pay["level"] = row.get("level", row.get("game", ""))
        pay["role"] = row.get("role", row.get("position", ""))
        pay["raw_line"] = json.dumps(row)
        
        return pay
        
    def _format_date(self, dt) -> str:
        if not dt:
            return ""
        if hasattr(dt, 'strftime'):
            return dt.strftime("%Y-%m-%d")
        
        dt_str = str(dt).strip()
        m = re.match(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4}|\d{2})", dt_str)
        if m:
            mm, dd, yy = m.groups()
            if len(yy) == 2:
                yy = "20" + yy
            return f"{int(yy):04d}-{int(mm):02d}-{int(dd):02d}"
        return dt_str
        
    def _build_result(self, payments: list[dict]) -> ParseResult:
        valid_payments = [p for p in payments if p["amount"] > 0]
        
        total_amount = sum(p["amount"] for p in valid_payments)
        preview = {
            "Total Games": len(valid_payments),
            "Total Earned": f"${total_amount:,.2f}"
        }
        
        if valid_payments:
            dates = [p["game_date"] for p in valid_payments if p["game_date"]]
            if dates:
                preview["Date Range"] = f"{min(dates)} to {max(dates)}"
                
        warnings = []
        can_commit = True
        if not valid_payments:
            warnings.append(
                "⚠ BLOCK: Recognized as an Eventlink export but no "
                "valid payments > $0 were extracted. The file format "
                "may have changed, or the export is empty."
            )
            can_commit = False

        return ParseResult(
            parser_type=self.parser_type,
            preview=preview,
            data={"payments": valid_payments},
            warnings=warnings,
            can_commit=can_commit,
        )

    def commit(self, conn: sqlite3.Connection, result: ParseResult) -> dict:
        """Write parsed data via dal.transactions.upsert_transactions.

        Routing through ``upsert_transactions`` enforces the canonical
        sign/direction invariant. The 7-day duplicate guard is a wider
        net than upsert's deterministic-hash dedup (which keys on exact
        posting_date), so we keep it as a pre-filter.
        """
        from dal.transactions import (
            upsert_transactions,
            derive_signed_amount,
            compute_txn_id,
        )

        payments = result.data.get("payments", [])

        if not payments:
            return {"inserted": 0, "duplicates": 0, "total_value": 0}

        acct_id = "eventlink_manual"

        conn.execute(
            "INSERT OR IGNORE INTO accounts (id, name, type, subtype, institution_name) VALUES (?, ?, 'depository', 'cash', 'Eventlink')",
            (acct_id, 'Eventlink Payouts')
        )

        txns: list[dict] = []
        duplicates = 0
        total_value = 0.0

        for pay in payments:
            amt = abs(float(pay["amount"]))
            pay_date = pay["pay_date"]

            parts = [p for p in [pay['sport'], pay['level'], pay['role']] if p]
            desc = f"Officiating: {' '.join(parts)}" if parts else "Eventlink Officiating"

            # Pre-upsert dedup: 7-day window + exact amount + Officiating
            # Income category. Wider than upsert_transactions' deterministic
            # hash, which would treat two payments to the same date+amount
            # as duplicates but two payments 5 days apart as distinct.
            dup_row = conn.execute("""
                SELECT id FROM transactions
                WHERE category = 'Officiating Income'
                  AND amount = ?
                  AND abs(julianday(posting_date) - julianday(?)) <= 7
                LIMIT 1
            """, (amt, pay_date)).fetchone()

            if dup_row:
                duplicates += 1
                continue

            txns.append({
                "account_id": acct_id,
                "institution_id": "eventlink",
                "posting_date": pay_date,
                "transaction_date": pay["game_date"],
                "amount": amt,
                "signed_amount": derive_signed_amount(amt, "Credit"),
                "direction": "Credit",
                "description": desc,
                "raw_description": pay["raw_line"],
                "category": "Officiating Income",
                "status": "posted",
            })
            total_value += amt

        stats = upsert_transactions(conn, txns)

        # `upsert_transactions` doesn't set the ``merchant`` column;
        # stamp it explicitly so the merchant trends report can pivot
        # on these rows.
        for txn in txns:
            txn_id = compute_txn_id(
                institution_id=txn["institution_id"],
                account_id=txn["account_id"],
                posting_date=txn["posting_date"],
                amount=txn["amount"],
                description=txn["description"],
                transaction_date=txn["transaction_date"],
                sequence_index=txn.get("sequence_index", 0),
            )
            conn.execute(
                "UPDATE transactions SET merchant = ? WHERE id = ?",
                ("Eventlink", txn_id),
            )

        return {
            "inserted": stats["inserted"],
            "duplicates_skipped": duplicates,
            "total_earned": total_value,
        }
